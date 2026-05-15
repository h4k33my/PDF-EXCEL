"""
Dialogs: import columns from another sheet, filter by numeric primary column.
"""
from __future__ import annotations

from typing import List, Literal, Optional, Tuple

from PyQt6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QComboBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QAbstractItemView,
    QDialogButtonBox,
    QLineEdit,
    QFileDialog,
    QRadioButton,
    QInputDialog,
    QHeaderView,
)
from PyQt6.QtGui import QColor
from openpyxl import load_workbook

from utils.templates import (
    delete_template as _delete_template,
    load_all_templates,
    session_to_template,
    template_to_session,
    upsert_template,
)


class ImportColumnsDialog(QDialog):
    def __init__(self, parent, sheets_data: List[dict]):
        super().__init__(parent)
        self.setWindowTitle("Import columns")
        self._sheets = sheets_data
        self._selected_source_cols: List[int] = []
        self._source_idx = 0
        self._dest_idx = 0

        layout = QVBoxLayout(self)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Source sheet:"))
        self._source_combo = QComboBox()
        for i, s in enumerate(sheets_data):
            self._source_combo.addItem(s["name"], i)
        self._source_combo.currentIndexChanged.connect(self._on_source_changed)
        row1.addWidget(self._source_combo)
        layout.addLayout(row1)

        self._preview = QTableWidget()
        self._preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._preview.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectColumns)
        self._preview.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._preview.horizontalHeader().setVisible(True)
        layout.addWidget(QLabel("Select columns (click column headers; Ctrl+click for multiple):"))
        layout.addWidget(self._preview, stretch=1)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Destination sheet:"))
        self._dest_combo = QComboBox()
        for i, s in enumerate(sheets_data):
            self._dest_combo.addItem(s["name"], i)
        self._dest_combo.addItem("New sheet (created on OK)", -1)
        row2.addWidget(self._dest_combo)
        layout.addLayout(row2)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._on_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self._reload_source_preview()
        self.resize(720, 420)

    def _on_source_changed(self, _index: int):
        self._reload_source_preview()

    def _reload_source_preview(self):
        idx = self._source_combo.currentData()
        if idx is None:
            idx = 0
        self._source_idx = idx
        data = self._sheets[self._source_idx]["data"]
        if not data:
            self._preview.clear()
            self._preview.setRowCount(0)
            self._preview.setColumnCount(0)
            return
        ncols = len(data[0])
        nrows = min(len(data), 500)
        self._preview.setColumnCount(ncols)
        self._preview.setRowCount(nrows)
        hdr_bg = QColor(68, 114, 196)
        hdr_fg = QColor(255, 255, 255)
        for r in range(nrows):
            row = data[r]
            for c in range(ncols):
                val = row[c] if c < len(row) else ""
                item = QTableWidgetItem(str(val))
                if r == 0:
                    item.setBackground(hdr_bg)
                    item.setForeground(hdr_fg)
                self._preview.setItem(r, c, item)
        self._preview.resizeColumnsToContents()

    def _on_ok(self):
        sel = self._preview.selectionModel().selectedColumns()
        cols = sorted({idx.column() for idx in sel})
        if not cols:
            QMessageBox.warning(self, "Import columns", "Select at least one column (use column headers).")
            return
        self._selected_source_cols = cols
        self._dest_idx = self._dest_combo.currentData()
        self.accept()

    def result(self) -> Tuple[int, List[int], int]:
        """source_sheet_idx, column_indices (sorted), dest_sheet_idx or -1 for new sheet."""
        return self._source_idx, list(self._selected_source_cols), self._dest_idx


class SheetSelectionDialog(QDialog):
    def __init__(self, parent, sheets_data: List[dict], *, title: str, prompt: str, default_idx: int = 0):
        super().__init__(parent)
        self.setWindowTitle(title)
        self._selected_idx: Optional[int] = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(prompt))
        self._combo = QComboBox()
        for i, s in enumerate(sheets_data):
            name = s.get("name", f"Sheet {i + 1}")
            self._combo.addItem(name, i)
        if 0 <= default_idx < self._combo.count():
            self._combo.setCurrentIndex(default_idx)
        layout.addWidget(self._combo)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._accept_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _accept_ok(self):
        self._selected_idx = self._combo.currentData()
        self.accept()

    def selected_sheet_index(self) -> Optional[int]:
        return self._selected_idx


class ColumnSelectionDialog(QDialog):
    def __init__(
        self,
        parent,
        header_row: List[object],
        *,
        title: str,
        prompt: str,
        default_idx: int = 0,
    ):
        super().__init__(parent)
        self.setWindowTitle(title)
        self._col_idx: Optional[int] = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(prompt))
        self._combo = QComboBox()
        for i, h in enumerate(header_row):
            label = str(h).strip() if h is not None else ""
            if not label:
                label = f"Column {i + 1}"
            self._combo.addItem(label, i)
        if 0 <= default_idx < self._combo.count():
            self._combo.setCurrentIndex(default_idx)
        layout.addWidget(self._combo)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._accept_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _accept_ok(self):
        self._col_idx = self._combo.currentData()
        self.accept()

    def selected_column_index(self) -> Optional[int]:
        return self._col_idx


class EventColumnModeDialog(QDialog):
    """Ask whether to insert a new Event column or use an existing sheet column."""

    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Event column setup")
        self._choice: Optional[Literal["create", "existing"]] = None

        layout = QVBoxLayout(self)
        layout.addWidget(
            QLabel(
                "How should the app set up the column where you choose an event for each row?"
            )
        )
        layout.addWidget(
            QLabel(
                "• Create new Event column: inserts a column titled “Event” to the right of a column you pick.\n"
                "• Use existing column: pick a column that is already in the sheet (header stays as-is)."
            )
        )
        row = QHBoxLayout()
        create_btn = QPushButton("Create new Event column")
        existing_btn = QPushButton("Use existing column")
        create_btn.clicked.connect(lambda: self._pick("create"))
        existing_btn.clicked.connect(lambda: self._pick("existing"))
        row.addWidget(create_btn)
        row.addWidget(existing_btn)
        layout.addLayout(row)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Cancel)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _pick(self, choice: Literal["create", "existing"]):
        self._choice = choice
        self.accept()

    def choice(self) -> Optional[Literal["create", "existing"]]:
        return self._choice


class UpdateExistingExcelDialog(QDialog):
    """Configure update mode for writing into an existing workbook."""

    def __init__(self, parent, *, active_sheet_name: str, has_selection: bool):
        super().__init__(parent)
        self.setWindowTitle("Update existing Excel workbook")
        self._result = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Destination workbook (.xlsx):"))
        path_row = QHBoxLayout()
        self._path_edit = QLineEdit()
        self._path_edit.setPlaceholderText("Choose existing workbook")
        browse_btn = QPushButton("Browse…")
        browse_btn.clicked.connect(self._browse_workbook)
        path_row.addWidget(self._path_edit, 1)
        path_row.addWidget(browse_btn)
        layout.addLayout(path_row)

        layout.addWidget(QLabel("Operation:"))
        self._mode_combo = QComboBox()
        self._mode_combo.addItem("Add selected app sheets as new sheet(s)", "add_sheets")
        self._mode_combo.addItem("Paste into existing sheet", "paste_range")
        self._mode_combo.currentIndexChanged.connect(self._sync_visibility)
        layout.addWidget(self._mode_combo)

        self._paste_block = QVBoxLayout()
        self._sheet_combo = QComboBox()
        self._start_cell_edit = QLineEdit("A1")
        self._start_cell_edit.setPlaceholderText("A1")
        self._source_combo = QComboBox()
        if has_selection:
            self._source_combo.addItem("Use highlighted cells on active sheet", "selection")
        self._source_combo.addItem(f"Use full active sheet ({active_sheet_name})", "full_active")
        self._paste_block.addWidget(QLabel("Destination sheet:"))
        self._paste_block.addWidget(self._sheet_combo)
        self._paste_block.addWidget(QLabel("Start cell (e.g. A2500):"))
        self._paste_block.addWidget(self._start_cell_edit)
        self._paste_block.addWidget(QLabel("Source to paste:"))
        self._paste_block.addWidget(self._source_combo)
        layout.addLayout(self._paste_block)

        self._clear_first = QRadioButton("Clear destination rectangle first, then paste values")
        self._clear_first.setChecked(False)
        layout.addWidget(self._clear_first)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._accept_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.resize(640, 320)
        self._sync_visibility()

    def _sync_visibility(self):
        paste_mode = self._mode_combo.currentData() == "paste_range"
        for i in range(self._paste_block.count()):
            item = self._paste_block.itemAt(i)
            w = item.widget()
            if w is not None:
                w.setVisible(paste_mode)
        self._clear_first.setVisible(paste_mode)

    def _browse_workbook(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select destination workbook", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        self._path_edit.setText(path)
        self._reload_sheet_names(path)

    def _reload_sheet_names(self, path: str):
        self._sheet_combo.clear()
        try:
            wb = load_workbook(path, read_only=True)
            for name in wb.sheetnames:
                self._sheet_combo.addItem(name, name)
            wb.close()
        except Exception:
            self._sheet_combo.clear()

    def _accept_ok(self):
        path = str(self._path_edit.text() or "").strip()
        if not path:
            QMessageBox.warning(self, "Update workbook", "Choose a destination workbook.")
            return
        mode = self._mode_combo.currentData()
        if mode == "paste_range":
            if self._sheet_combo.count() == 0:
                QMessageBox.warning(self, "Update workbook", "No destination sheets found in workbook.")
                return
            sheet_name = self._sheet_combo.currentData()
            start_cell = str(self._start_cell_edit.text() or "").strip().upper()
            if not start_cell or not any(ch.isdigit() for ch in start_cell):
                QMessageBox.warning(self, "Update workbook", "Enter a valid start cell like A1.")
                return
            source_mode = self._source_combo.currentData()
            self._result = {
                "path": path,
                "mode": mode,
                "sheet_name": sheet_name,
                "start_cell": start_cell,
                "source_mode": source_mode,
                "clear_first": bool(self._clear_first.isChecked()),
            }
        else:
            self._result = {
                "path": path,
                "mode": mode,
            }
        self.accept()

    def result(self) -> Optional[dict]:
        return self._result


class EventTemplateDialog(QDialog):
    """Manage saved event-header templates: load, save, edit headers + aliases, delete."""

    def __init__(self, parent, *, current_options: List[str], current_aliases: dict):
        super().__init__(parent)
        self.setWindowTitle("Event header templates")
        self._loaded_options: Optional[List[str]] = None
        self._loaded_aliases: Optional[dict] = None

        layout = QVBoxLayout(self)

        layout.addWidget(QLabel("Saved templates:"))
        picker_row = QHBoxLayout()
        self._template_combo = QComboBox()
        self._template_combo.currentIndexChanged.connect(self._on_template_picked)
        picker_row.addWidget(self._template_combo, 1)
        new_btn = QPushButton("New")
        new_btn.setToolTip("Start with the current session's headers (or empty)")
        new_btn.clicked.connect(self._on_new_clicked)
        delete_btn = QPushButton("Delete")
        delete_btn.clicked.connect(self._on_delete_clicked)
        picker_row.addWidget(new_btn)
        picker_row.addWidget(delete_btn)
        layout.addLayout(picker_row)

        layout.addWidget(QLabel(
            "Headers and their alias keywords (comma-separated). The matcher checks "
            "if any header-name word OR any alias appears in a transaction's description."
        ))
        self._table = QTableWidget(0, 2)
        self._table.setHorizontalHeaderLabels(["Header name", "Aliases (comma-separated)"])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._table.verticalHeader().setVisible(False)
        layout.addWidget(self._table, stretch=1)

        row_btn_row = QHBoxLayout()
        add_row_btn = QPushButton("Add header")
        add_row_btn.clicked.connect(self._add_blank_row)
        del_row_btn = QPushButton("Remove selected")
        del_row_btn.clicked.connect(self._remove_selected_rows)
        row_btn_row.addWidget(add_row_btn)
        row_btn_row.addWidget(del_row_btn)
        row_btn_row.addStretch()
        layout.addLayout(row_btn_row)

        action_row = QHBoxLayout()
        save_btn = QPushButton("Save as…")
        save_btn.clicked.connect(self._on_save_as)
        load_btn = QPushButton("Load into session")
        load_btn.clicked.connect(self._on_load_into_session)
        cancel_btn = QPushButton("Close")
        cancel_btn.clicked.connect(self.reject)
        action_row.addWidget(save_btn)
        action_row.addWidget(load_btn)
        action_row.addStretch()
        action_row.addWidget(cancel_btn)
        layout.addLayout(action_row)

        self.resize(640, 460)
        self._reload_template_list(select_name=None)
        # Pre-fill the editor with the current session's options/aliases
        self._populate_editor(current_options, current_aliases)

    def _reload_template_list(self, *, select_name: Optional[str]):
        self._template_combo.blockSignals(True)
        try:
            self._template_combo.clear()
            self._template_combo.addItem("(unsaved)", None)
            templates = load_all_templates()
            for t in templates:
                name = str(t.get("name", "")).strip()
                if not name:
                    continue
                self._template_combo.addItem(name, name)
            if select_name:
                idx = self._template_combo.findData(select_name)
                if idx >= 0:
                    self._template_combo.setCurrentIndex(idx)
        finally:
            self._template_combo.blockSignals(False)

    def _populate_editor(self, options: List[str], aliases: dict):
        self._table.setRowCount(0)
        for opt in options:
            self._add_row(opt, ", ".join(aliases.get(opt, []) or []))
        if not options:
            self._add_blank_row()

    def _add_row(self, name: str, aliases_csv: str):
        r = self._table.rowCount()
        self._table.insertRow(r)
        self._table.setItem(r, 0, QTableWidgetItem(str(name or "")))
        self._table.setItem(r, 1, QTableWidgetItem(str(aliases_csv or "")))

    def _add_blank_row(self):
        self._add_row("", "")

    def _remove_selected_rows(self):
        rows = sorted({i.row() for i in self._table.selectedIndexes()}, reverse=True)
        for r in rows:
            self._table.removeRow(r)

    def _on_template_picked(self, _index: int):
        name = self._template_combo.currentData()
        if not name:
            return
        for t in load_all_templates():
            if str(t.get("name", "")).strip() == name:
                opts, aliases = template_to_session(t)
                self._populate_editor(opts, aliases)
                return

    def _on_new_clicked(self):
        self._template_combo.setCurrentIndex(0)
        self._table.setRowCount(0)
        self._add_blank_row()

    def _on_delete_clicked(self):
        name = self._template_combo.currentData()
        if not name:
            QMessageBox.information(self, "Delete template", "Pick a saved template first.")
            return
        if QMessageBox.question(
            self, "Delete template", f"Delete template “{name}”?"
        ) != QMessageBox.StandardButton.Yes:
            return
        _delete_template(name)
        self._reload_template_list(select_name=None)

    def _collect_editor_state(self) -> tuple[List[str], dict]:
        options: List[str] = []
        aliases: dict = {}
        for r in range(self._table.rowCount()):
            name_item = self._table.item(r, 0)
            alias_item = self._table.item(r, 1)
            name = (name_item.text() if name_item else "").strip()
            if not name:
                continue
            if name in options:
                continue
            options.append(name)
            alias_csv = alias_item.text() if alias_item else ""
            alias_list = [a.strip() for a in alias_csv.split(",") if a.strip()]
            if alias_list:
                aliases[name] = alias_list
        return options, aliases

    def _on_save_as(self):
        options, aliases = self._collect_editor_state()
        if not options:
            QMessageBox.warning(self, "Save template", "Add at least one header before saving.")
            return
        suggested = self._template_combo.currentData() or ""
        name, ok = QInputDialog.getText(self, "Save template", "Template name:", text=suggested)
        if not ok:
            return
        name = (name or "").strip()
        if not name:
            QMessageBox.warning(self, "Save template", "Template name cannot be empty.")
            return
        try:
            template = session_to_template(name, options, aliases)
            upsert_template(name, template["headers"])
        except Exception as e:
            QMessageBox.critical(self, "Save template", f"Could not save: {e}")
            return
        self._reload_template_list(select_name=name)

    def _on_load_into_session(self):
        options, aliases = self._collect_editor_state()
        if not options:
            QMessageBox.warning(self, "Load template", "Add at least one header before loading.")
            return
        self._loaded_options = options
        self._loaded_aliases = aliases
        self.accept()

    def loaded_session(self) -> Optional[tuple]:
        """Returns (options, aliases) if 'Load into session' was clicked; else None."""
        if self._loaded_options is None:
            return None
        return self._loaded_options, self._loaded_aliases or {}
