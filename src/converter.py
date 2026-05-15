"""
Verbatim PDF-to-Excel converter for bank statements.

Extracts each table block from a PDF as-is, preserving original column count
and header wording. Distinct account blocks (bank-details + transaction tables)
become separate sheets. No canonicalization, no synthetic period rows.

Falls back to coordinate-based reconstruction for borderless PDFs where
pdfplumber's table extraction yields too little content.
"""
import re
import pdfplumber
from coordinate_fallback import reconstruct_transactions_coordinate, should_use_coordinate_fallback
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

PAGE_MARKER_RE = re.compile(r"\b(?:page|pg|p|pag)\.?\s*\d+\b", re.I)
INLINE_PAGE_ARTIFACT_RE = re.compile(r"\b(?:pag|page)\b\s*\d*\b", re.I)
PAGE_FRAGMENT_RE = re.compile(r"\s*\b[ep]\s+\d+\b\s*", re.I)  # Catches " e 1", " p 2", " Page 1" fragments
NUMERIC_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?$|^-?\d+(?:\.\d{1,2})?$")
DATE_RE = re.compile(
    r"^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+\d{2,4}|\d{1,2}-(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{2,4})$",
    re.I,
)

TRANSACTION_HEADER_KEYWORDS = (
    'transaction date', 'book date', 'tran date', 'value date', 'date',
    'details', 'description', 'narration', 'particulars', 'remarks',
    'debit', 'withdrawal', 'dr', 'credit', 'deposit', 'cr', 'lodgement',
    'balance', 'amount', 'reference', 'ref',
)

BANK_DETAIL_KEYWORDS = (
    'account number', 'account no', 'account name', 'account type', 'currency',
    'opening balance', 'closing balance', 'usable balance', 'available balance',
    'branch', 'address', 'statement period', 'print date', 'print. date',
    'total debit', 'total credit', 'internal reference', 'iban',
    'account summary', 'short name',
)


def clean_cell(cell):
    """Remove newlines, page markers, and normalize whitespace."""
    if cell is None:
        return ""
    s = str(cell).replace('\n', ' ').strip()
    s = PAGE_MARKER_RE.sub('', s).strip()
    s = INLINE_PAGE_ARTIFACT_RE.sub('', s).strip()
    s = PAGE_FRAGMENT_RE.sub('', s).strip()  # Remove page number fragments like " e 1", " p 2"
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def is_date_like(value):
    """Return True if value resembles a statement date."""
    if value is None:
        return False
    return bool(DATE_RE.match(str(value).strip()))


def to_float(value):
    """Parse numeric values with thousand separators."""
    if value is None:
        return None
    s = str(value).strip().replace(',', '')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def first_numeric_token(value):
    """Extract first number token from mixed artifacts like '74.26 995,412.22'."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if NUMERIC_RE.match(s):
        return s
    for token in re.split(r"\s+", s):
        token = token.strip()
        if NUMERIC_RE.match(token):
            return token
    return s


def is_page_marker_row(row):
    """Return True if a row contains only page-number markers."""
    if not row:
        return False
    non_empty = [cell.strip() for cell in row if cell and cell.strip()]
    if not non_empty:
        return False
    return all(PAGE_MARKER_RE.fullmatch(cell) for cell in non_empty)


def drop_empty_columns(table_rows):
    """Remove columns that are blank in every row."""
    if not table_rows:
        return []
    max_cols = max(len(row) for row in table_rows)
    normalized = [row + [""] * (max_cols - len(row)) for row in table_rows]
    keep_columns = [
        col_idx for col_idx in range(max_cols)
        if any(normalized[row_idx][col_idx].strip() for row_idx in range(len(normalized)))
    ]
    if not keep_columns:
        return []
    return [[row[col_idx] for col_idx in keep_columns] for row in normalized]


def is_continuation_row(row):
    """Check if a row is a continuation (multi-line description spillover)."""
    if not row:
        return True
    first_cell = str(row[0]).strip() if row[0] else ""
    row_text = ' '.join(str(c).lower() for c in row if c)
    # Standalone summary rows (Total / Closing Balance / etc.) are not continuations.
    if any(k in row_text for k in ('total', 'closing balance', 'balance b/f', 'grand total', 'balance as at')):
        return False
    # Transaction header rows are not continuations (would otherwise drop the header).
    header_hits = sum(1 for k in TRANSACTION_HEADER_KEYWORDS if k in row_text)
    if header_hits >= 3 and not is_date_like(first_cell):
        return False
    # Rows with empty first cell are continuations only if they have no numeric content
    # (an empty-date row carrying amount/balance values is a standalone record, not overflow).
    if not first_cell:
        for cell in row[1:]:
            cell_str = str(cell).strip() if cell else ""
            if cell_str and to_float(cell_str) is not None:
                return False
        return True
    return not bool(DATE_RE.match(first_cell))


def merge_continuation_rows(table_rows):
    """Merge rows that are continuations of previous transactions."""
    if not table_rows:
        return []
    merged = []
    current_row = None
    for row in table_rows:
        if is_continuation_row(row):
            if current_row is not None:
                for i, cell in enumerate(row):
                    cell_str = str(cell).strip() if cell else ""
                    if cell_str and i < len(current_row):
                        current_part = str(current_row[i]).strip() if current_row[i] else ""
                        current_row[i] = f"{current_part} {cell_str}".strip()
        else:
            if current_row is not None:
                merged.append(current_row)
            current_row = [clean_cell(cell) for cell in row]
    if current_row is not None:
        merged.append(current_row)
    return merged


def _clean_table_rows(raw_rows):
    """Clean cells; drop page-marker rows and fully empty rows. No continuation merging."""
    cleaned = []
    for row in raw_rows:
        if not row:
            continue
        cleaned_row = [clean_cell(cell) for cell in row]
        if is_page_marker_row(cleaned_row):
            continue
        if not any(cell.strip() for cell in cleaned_row):
            continue
        cleaned.append(cleaned_row)
    return cleaned


def _table_text_blob(table_rows):
    return ' '.join(str(c).lower() for row in table_rows for c in row if c)


def _detect_transaction_header_index(table_rows):
    """Find header row index in a transaction table; None if absent."""
    for idx, row in enumerate(table_rows[:5]):
        row_text = ' '.join(str(c).lower() for c in row if c)
        hits = sum(1 for k in TRANSACTION_HEADER_KEYWORDS if k in row_text)
        if hits >= 3:
            first = str(row[0]).strip() if row else ''
            if not is_date_like(first):
                return idx
    return None


def _classify_table(normalized_rows):
    """Return 'bank_details', 'transactions', or 'other'."""
    if not normalized_rows:
        return 'other'
    text = _table_text_blob(normalized_rows)
    max_cols = max(len(r) for r in normalized_rows)

    # Bank details: vertical key/value layout (1-3 cols), bank-info keywords.
    if max_cols <= 3:
        bd_hits = sum(1 for k in BANK_DETAIL_KEYWORDS if k in text)
        if bd_hits >= 2:
            return 'bank_details'

    # Transactions: 4+ cols with header-row keywords or date-pattern data rows.
    if max_cols >= 4:
        if _detect_transaction_header_index(normalized_rows) is not None:
            return 'transactions'
        date_rows = sum(
            1 for row in normalized_rows[:10]
            if row and DATE_RE.match(str(row[0]).strip())
        )
        if date_rows >= 2:
            return 'transactions'

    return 'other'


def _headers_match(h1, h2):
    """Compare two header rows by their non-empty cells (case-insensitive)."""
    n1 = [str(c).strip().lower() for c in h1 if str(c).strip()]
    n2 = [str(c).strip().lower() for c in h2 if str(c).strip()]
    return n1 == n2


def _bank_details_lookup(bank_details_rows):
    """Build {label_lower: value} from key/value bank-details rows."""
    lookup = {}
    for row in bank_details_rows:
        if len(row) < 2:
            continue
        label = str(row[0]).strip().lower()
        value = str(row[1]).strip()
        if label and value and label not in lookup:
            lookup[label] = value
    return lookup


def _short_currency(value):
    """Map full currency text to a short code where possible."""
    v = (value or "").upper()
    if 'NAIRA' in v:
        return 'NGN'
    if 'DOLLAR' in v:
        return 'USD'
    if 'POUND' in v or 'STERLING' in v:
        return 'GBP'
    if 'EURO' in v:
        return 'EUR'
    if 'YUAN' in v or 'RENMINBI' in v:
        return 'CNY'
    return value[:6].strip()


def _derive_sheet_name(bank_details_rows, fallback):
    """Derive a sheet name like 'NGN - 0571440174' from bank details if possible."""
    if not bank_details_rows:
        return fallback
    lookup = _bank_details_lookup(bank_details_rows)
    currency = ''
    account_no = ''
    for k, v in lookup.items():
        if 'currency' in k and not currency:
            currency = _short_currency(v)
        if ('account' in k and ('no' in k or 'number' in k)) and not account_no:
            account_no = v
    parts = [p for p in [currency, account_no] if p]
    return ' - '.join(parts) if parts else fallback


def _make_unique_sheet_name(name, used):
    """Append _2, _3, etc. if name already in `used`. Mutates `used`."""
    if name not in used:
        used.add(name)
        return name
    n = 2
    while f"{name}_{n}" in used:
        n += 1
    new_name = f"{name}_{n}"
    used.add(new_name)
    return new_name


def _build_generic_table_sheet(table_rows, used_names, index):
    """Build a generic sheet from a raw table that isn't bank details or transactions."""
    if not table_rows:
        return None
    max_cols = max(len(r) for r in table_rows)
    sheet_data = [r + [''] * (max_cols - len(r)) for r in table_rows]
    sheet_data = drop_empty_columns(sheet_data)
    if not sheet_data:
        return None
    header_row = table_rows[0] if table_rows else []
    header_text = ' '.join(str(c).strip() for c in header_row if c).strip()
    fallback = header_text or f"Table {index}"
    base = _sanitize_sheet_name(fallback, used_names)
    name = _make_unique_sheet_name(base, used_names)
    return {'name': name, 'data': sheet_data, 'is_table': True}


def _build_full_text_sheet(pdf_path, used_names):
    """Extract full text from PDF and build a sheet with it."""
    text_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            if text.strip():
                text_data.append([f"Page {page_num}"])
                for line in text.split('\n'):
                    if line.strip():
                        text_data.append([line.strip()])
                text_data.append([])  # Blank row between pages
    if not text_data:
        return None
    base = "Full Text"
    name = _make_unique_sheet_name(base, used_names)
    return {'name': name, 'data': text_data, 'is_table': False}


def _build_sheet_from_block(block, used_names, total_blocks):
    """Combine bank_details + transactions into one verbatim sheet."""
    sheet_data = []
    if block['bank_details']:
        sheet_data.extend(block['bank_details'])
    if block['transactions']:
        sheet_data.extend(block['transactions'])
    if not sheet_data:
        return None
    max_cols = max(len(r) for r in sheet_data)
    sheet_data = [r + [''] * (max_cols - len(r)) for r in sheet_data]
    sheet_data = drop_empty_columns(sheet_data)
    if not sheet_data:
        return None
    fallback = 'Statement' if total_blocks == 1 else f"Statement_{len(used_names) + 1}"
    base = _derive_sheet_name(block['bank_details'], fallback)
    name = _make_unique_sheet_name(base, used_names)
    return {'name': name, 'data': sheet_data, 'is_table': True}


def _coord_fallback_sheet(pdf_path, blocks):
    """Try coordinate-based extraction; build a single sheet if it yields data."""
    try:
        coord_headers, coord_rows = reconstruct_transactions_coordinate(pdf_path)
    except Exception:
        return None
    if not coord_rows:
        return None
    sheet_data = []
    # Preserve any bank details from the first detected block
    if blocks and blocks[0]['bank_details']:
        sheet_data.extend(blocks[0]['bank_details'])
    if coord_headers:
        sheet_data.append(list(coord_headers))
    sheet_data.extend(coord_rows)
    max_cols = max(len(r) for r in sheet_data) if sheet_data else 0
    sheet_data = [r + [''] * (max_cols - len(r)) for r in sheet_data]
    sheet_data = drop_empty_columns(sheet_data)
    if not sheet_data:
        return None
    base = _derive_sheet_name(blocks[0]['bank_details'] if blocks else [], 'Statement')
    return {'name': base, 'data': sheet_data, 'is_table': True}


def extract_all_tables_from_pdf(pdf_path):
    """
    Extract a bank-statement PDF verbatim. Returns a list of sheets, one per
    detected account block. Preserves original column count and header
    wording exactly. Falls back to coordinate extraction for weak PDFs.
    """
    blocks = []  # [{'bank_details': [rows], 'transactions': [rows], 'last_tx_header': [row]|None}]

    with pdfplumber.open(pdf_path) as pdf:
        current = {'bank_details': [], 'transactions': [], 'last_tx_header': None}
        other_tables = []

        for page in pdf.pages:
            page_tables = page.extract_tables() or []
            for raw_table in page_tables:
                if not raw_table:
                    continue
                cleaned = _clean_table_rows(raw_table)
                if not cleaned:
                    continue
                kind = _classify_table(cleaned)

                if kind == 'bank_details':
                    # Vertical key/value layout — keep as-is, never merge continuations.
                    if current['transactions'] or current['bank_details']:
                        blocks.append(current)
                        current = {'bank_details': [], 'transactions': [], 'last_tx_header': None}
                    current['bank_details'] = cleaned
                elif kind == 'transactions':
                    # Transaction tables can have multi-line descriptions; merge those.
                    merged = merge_continuation_rows(cleaned)
                    hdr_idx = _detect_transaction_header_index(merged)
                    if hdr_idx is not None:
                        header_row = merged[hdr_idx]
                        if current['last_tx_header'] is None:
                            current['transactions'].extend(merged[hdr_idx:])
                            current['last_tx_header'] = header_row
                        elif _headers_match(current['last_tx_header'], header_row):
                            current['transactions'].extend(merged[hdr_idx + 1:])
                        else:
                            current['transactions'].extend(merged[hdr_idx:])
                            current['last_tx_header'] = header_row
                    else:
                        current['transactions'].extend(merged)
                else:
                    other_tables.append(cleaned)

        if current['bank_details'] or current['transactions']:
            blocks.append(current)

    # Coordinate fallback when pdfplumber yielded no transactions or extraction looks weak.
    primary_tx_count = sum(
        max(0, len(b['transactions']) - (1 if b['last_tx_header'] is not None else 0))
        for b in blocks
    )
    needs_fallback = (
        primary_tx_count == 0
        or should_use_coordinate_fallback(pdf_path, primary_tx_count)
    )
    if needs_fallback:
        coord_sheet = _coord_fallback_sheet(pdf_path, blocks)
        if coord_sheet is not None:
            return [coord_sheet]

    sheets = []
    used_names = set()
    for block in blocks:
        sheet = _build_sheet_from_block(block, used_names, total_blocks=len(blocks))
        if sheet is not None:
            sheets.append(sheet)
    for idx, table in enumerate(other_tables, start=1):
        sheet = _build_generic_table_sheet(table, used_names, idx)
        if sheet is not None:
            sheets.append(sheet)
    # Always add full text sheet if there's any text
    text_sheet = _build_full_text_sheet(pdf_path, used_names)
    if text_sheet is not None:
        sheets.append(text_sheet)
    return sheets


def export_to_excel(sheets_data, output_path):
    """Write sheets to Excel file with basic formatting."""
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for sheet_info in sheets_data:
        sheet_name = sheet_info['name']
        data = sheet_info['data']
        sheet_name = sheet_name[:31].replace('[', '').replace(']', '').replace(':', '').replace('*', '').replace('?', '').replace('/', '')
        ws = wb.create_sheet(sheet_name)
        for row_idx, row in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    wb.save(output_path)


def _sanitize_sheet_name(name, existing_names):
    base = str(name or "Sheet").strip() or "Sheet"
    cleaned = base[:31].replace('[', '').replace(']', '').replace(':', '').replace('*', '').replace('?', '').replace('/', '')
    if not cleaned:
        cleaned = "Sheet"
    candidate = cleaned
    n = 2
    existing_lower = {str(v).lower() for v in existing_names}
    while candidate.lower() in existing_lower:
        suffix = f"_{n}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        n += 1
    return candidate


def append_sheets_to_existing_workbook(dest_path, sheets_data):
    """Append app sheets as new sheets to an existing workbook (value-only writes)."""
    wb = load_workbook(dest_path)
    existing_names = list(wb.sheetnames)
    for sheet_info in sheets_data:
        src_name = sheet_info.get("name", "Sheet")
        data = sheet_info.get("data", [])
        safe_name = _sanitize_sheet_name(src_name, existing_names)
        ws = wb.create_sheet(safe_name)
        existing_names.append(safe_name)
        for r, row in enumerate(data, 1):
            for c, value in enumerate(row, 1):
                ws.cell(row=r, column=c).value = value
    wb.save(dest_path)
    wb.close()


def paste_values_into_existing_sheet(dest_path, sheet_name, start_cell, grid, *, clear_grid=False):
    """
    Paste grid values into existing worksheet from start_cell, preserving destination formatting.
    Optionally clear the whole destination rectangle before writing.
    """
    wb = load_workbook(dest_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Destination sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]
    start_row, start_col = coordinate_to_tuple(start_cell)
    rows = len(grid)
    cols = max((len(r) for r in grid), default=0)
    if clear_grid and rows > 0 and cols > 0:
        for rr in range(rows):
            for cc in range(cols):
                ws.cell(row=start_row + rr, column=start_col + cc).value = None
    for rr, row in enumerate(grid):
        for cc in range(cols):
            value = row[cc] if cc < len(row) else ""
            ws.cell(row=start_row + rr, column=start_col + cc).value = value
    wb.save(dest_path)
    wb.close()


def has_nonempty_cells_in_target_range(dest_path, sheet_name, start_cell, rows, cols):
    """Return True when destination range already has non-empty cell values."""
    if rows <= 0 or cols <= 0:
        return False
    wb = load_workbook(dest_path, read_only=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Destination sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]
    start_row, start_col = coordinate_to_tuple(start_cell)
    try:
        for rr in range(rows):
            for cc in range(cols):
                value = ws.cell(row=start_row + rr, column=start_col + cc).value
                if value not in (None, ""):
                    return True
        return False
    finally:
        wb.close()
