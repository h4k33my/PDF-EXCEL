"""
Load Excel workbooks into the same sheet dict structure used by the PDF converter.

Supports: .xlsx, .xlsm (via openpyxl), legacy .xls (via xlrd), and HTML-as-.xls
(banks commonly export HTML tables with a .xls extension).
"""
from __future__ import annotations

from typing import Any, List
import os


def _cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def _load_html_as_sheets(path: str) -> List[dict]:
    """Parse an HTML file (often disguised as .xls) and extract <table> elements as sheets."""
    from html.parser import HTMLParser

    class _TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.tables: list = []
            self._tbl: list | None = None
            self._row: list | None = None
            self._cell: list | None = None

        def handle_starttag(self, tag, attrs):
            t = tag.lower()
            if t == "table":
                self._tbl = []
            elif t == "tr" and self._tbl is not None:
                self._row = []
            elif t in ("td", "th") and self._row is not None:
                self._cell = []
            elif t == "br" and self._cell is not None:
                self._cell.append(" ")

        def handle_endtag(self, tag):
            t = tag.lower()
            if t == "table" and self._tbl is not None:
                self.tables.append(self._tbl)
                self._tbl = None
                self._row = None
            elif t == "tr" and self._row is not None:
                if any(c.strip() for c in self._row):
                    if self._tbl is not None:
                        self._tbl.append(self._row)
                self._row = None
            elif t in ("td", "th") and self._cell is not None:
                if self._row is not None:
                    self._row.append("".join(self._cell).replace("\xa0", " ").strip())
                self._cell = None

        def handle_data(self, data):
            if self._cell is not None:
                self._cell.append(data)

    content = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                content = f.read()
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if content is None:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()

    p = _TableParser()
    p.feed(content)

    out: List[dict] = []
    for i, tbl in enumerate(p.tables):
        if not tbl:
            continue
        max_cols = max(len(r) for r in tbl)
        rows = [r + [""] * (max_cols - len(r)) for r in tbl]
        while rows and not any(c.strip() for c in rows[-1]):
            rows.pop()
        if not rows:
            continue
        out.append({"name": f"Sheet{i + 1}", "data": rows, "is_table": True})

    return out or [{"name": "Sheet1", "data": [[""]], "is_table": True}]


def load_xlsx_to_sheets_data(path: str) -> List[dict]:
    """
    Read all worksheets from an Excel file (.xlsx, .xlsm, .xls).
    Returns [{'name': str, 'data': list[list[str]], 'is_table': True}, ...]
    """
    ext = os.path.splitext(path)[1].lower()
    out: List[dict] = []
    if ext in ('.xlsx', '.xlsm'):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([_cell(c) for c in row])
                # Trim trailing completely empty rows
                while rows and not any(str(c).strip() for c in rows[-1]):
                    rows.pop()
                if not rows:
                    rows = [[""]]
                # Normalize row widths
                max_cols = max(len(r) for r in rows) if rows else 1
                rows = [r + [""] * (max_cols - len(r)) for r in rows]
                name = ws.title[:31] or "Sheet"
                out.append({"name": name, "data": rows, "is_table": True})
        finally:
            wb.close()
        return out

    if ext == '.xls':
        # Sniff for HTML-as-.xls (banks commonly export HTML tables with .xls extension)
        with open(path, "rb") as f:
            raw_head = f.read(32)
        stripped = raw_head.lstrip(b"\r\n \t")
        # Strip UTF-8 BOM if present
        if stripped.startswith(b"\xef\xbb\xbf"):
            stripped = stripped[3:]
        if stripped[:1] == b"<":
            return _load_html_as_sheets(path)

        # xlrd is used for genuine legacy .xls binary files
        try:
            import xlrd  # type: ignore
        except Exception as e:
            raise ImportError("xlrd is required to read .xls files; please install it (pip install xlrd)") from e

        try:
            wb = xlrd.open_workbook(path, formatting_info=False)
        except Exception:
            # Last-resort fallback: file may still be HTML despite non-HTML header bytes
            return _load_html_as_sheets(path)

        for sheet in wb.sheets():
            rows = []
            for ri in range(sheet.nrows):
                row = sheet.row_values(ri)
                rows.append([_cell(c) for c in row])
            # Trim trailing completely empty rows
            while rows and not any(str(c).strip() for c in rows[-1]):
                rows.pop()
            if not rows:
                rows = [[""]]
            max_cols = max(len(r) for r in rows) if rows else 1
            rows = [r + [""] * (max_cols - len(r)) for r in rows]
            name = (sheet.name[:31] or "Sheet")
            out.append({"name": name, "data": rows, "is_table": True})
        return out

    raise ValueError(f"Unsupported Excel extension: {ext}")
