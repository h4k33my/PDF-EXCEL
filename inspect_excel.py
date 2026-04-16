#!/usr/bin/env python3
"""
Inspect the generated Excel file to verify correct table structure
"""
import openpyxl
import os

def inspect_excel(file_path):
    print(f"Inspecting Excel file: {file_path}")

    if not os.path.exists(file_path):
        print("File does not exist!")
        return

    wb = openpyxl.load_workbook(file_path)

    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")

        # Print first 10 rows
        print("First 10 rows:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_values.append(str(cell_value or "")[:50])
            print(f"Row {row_idx}: {' | '.join(row_values)}")

        # Check for empty rows or columns
        empty_rows = 0
        for row_idx in range(1, ws.max_row + 1):
            row_empty = True
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=row_idx, column=col_idx).value:
                    row_empty = False
                    break
            if row_empty:
                empty_rows += 1

        print(f"Empty rows: {empty_rows}")
        print(f"Data rows: {ws.max_row - empty_rows}")

if __name__ == "__main__":
    inspect_excel(r"c:\Users\DELL\Desktop\ZAPPHAIRE_TEST_CONVERSION_OUTPUT.xlsx")